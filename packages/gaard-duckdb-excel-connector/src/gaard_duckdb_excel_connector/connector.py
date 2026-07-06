from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from itertools import chain
import os
from pathlib import Path
import re
import sys
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlencode

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Connection
from sqlalchemy.exc import SQLAlchemyError

from gaard_core.schema.models import ColumnInfo, DatabaseSchema, TableInfo
from gaard_connectors.registry import ConnectorDefinition, ConnectorRegistry, ConnectorRegistryError
from gaard_connectors.sqlalchemy.executor import SQLAlchemyQueryExecutor
from gaard_plugin_api import ExtensionContext


CONNECTOR_TYPE_KEY = "duckdb-excel"
DUCKDB_EXCEL_URL_PREFIX = "duckdb-excel://"
RELATION_NAME_PATTERN = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]{0,62}$")
RESERVED_RELATION_NAMES = {
    "all",
    "and",
    "as",
    "by",
    "create",
    "delete",
    "describe",
    "drop",
    "from",
    "group",
    "insert",
    "join",
    "limit",
    "not",
    "or",
    "order",
    "select",
    "table",
    "update",
    "view",
    "where",
}
DEFAULT_MODE = "table"


@dataclass(frozen=True, slots=True)
class ExcelSheetMapping:
    sheet_name: str
    relation_name: str
    header: bool = True
    range: str | None = None
    all_varchar: bool = False
    ignore_errors: bool = False
    type_sample_rows: int | None = None


@dataclass(frozen=True, slots=True)
class DuckDBExcelConfig:
    catalog_path: Path
    workbook_path: Path
    sheets: tuple[ExcelSheetMapping, ...]
    always_read_current_file: bool = False

    @property
    def duckdb_url(self) -> str:
        return f"duckdb:///{self.catalog_path.as_posix()}"


def register(context: ExtensionContext) -> None:
    if not isinstance(context.registry, ConnectorRegistry):
        raise ConnectorRegistryError("DuckDB Excel connector requires a ConnectorRegistry.")

    context.registry.register(create_connector_definition())


def create_connector_definition() -> ConnectorDefinition:
    return ConnectorDefinition(
        type_key=CONNECTOR_TYPE_KEY,
        label="Excel File Loader",
        description=(
            "Loads configured Excel workbook sheets and exposes them as ordinary "
            "queryable datasource relations."
        ),
        sql_dialects=("duckdb",),
        url_prefixes=(DUCKDB_EXCEL_URL_PREFIX,),
        executor_factory=lambda database_url, max_rows: DuckDBExcelQueryExecutor(database_url, max_rows),
        introspector_factory=lambda database_url: DuckDBExcelSchemaIntrospector(database_url),
        connection_tester=test_connection,
        config_schema={
            "type": "object",
            "properties": {
                "database_url": {
                    "type": "string",
                    "title": "DuckDB Excel URL",
                    "description": (
                        "Use duckdb-excel:///path/source.xlsx for automatic sheet mapping, "
                        "or duckdb-excel:///catalog.duckdb?workbook=/path/source.xlsx"
                        "&sheet=Sheet:relation&mode=table for explicit mapping."
                    ),
                }
            },
            "required": ["database_url"],
        },
    )


class DuckDBExcelQueryExecutor(SQLAlchemyQueryExecutor):
    def __init__(self, database_url: str, max_rows: int = 100) -> None:
        config = parse_duckdb_excel_url(database_url)
        provision_excel_relations(config)
        super().__init__(database_url=config.duckdb_url, max_rows=max_rows)


class DuckDBExcelSchemaIntrospector:
    def __init__(self, database_url: str) -> None:
        self.config = parse_duckdb_excel_url(database_url)

    def introspect(self) -> DatabaseSchema:
        provision_excel_relations(self.config)
        engine = create_engine(self.config.duckdb_url)
        try:
            with engine.connect() as connection:
                return _introspect_configured_relations(connection, self.config)
        except SQLAlchemyError as exc:
            raise ValueError(_friendly_duckdb_error(exc)) from exc
        finally:
            engine.dispose()


def test_connection(database_url: str) -> None:
    config = parse_duckdb_excel_url(database_url)
    provision_excel_relations(config)

    engine = create_engine(config.duckdb_url)
    try:
        with engine.connect() as connection:
            connection.execute(text("LOAD excel"))
            for sheet in config.sheets:
                connection.execute(text(f'DESCRIBE "{sheet.relation_name}"'))
    finally:
        engine.dispose()


def build_duckdb_excel_url(
    catalog_path: Path,
    workbook_path: Path,
    sheets: list[tuple[str, str]],
    *,
    always_read_current_file: bool = False,
) -> str:
    query: list[tuple[str, str]] = [
        ("workbook", str(workbook_path)),
        ("mode", "view" if always_read_current_file else "table"),
    ]
    query.extend(("sheet", f"{sheet_name}:{relation_name}") for sheet_name, relation_name in sheets)
    return f"{DUCKDB_EXCEL_URL_PREFIX}{quote(str(catalog_path), safe=':/')}?{urlencode(query)}"


def parse_duckdb_excel_url(database_url: str) -> DuckDBExcelConfig:
    if not database_url.startswith(DUCKDB_EXCEL_URL_PREFIX):
        raise ValueError(f"DuckDB Excel URL must start with {DUCKDB_EXCEL_URL_PREFIX}.")

    raw_target = database_url.removeprefix(DUCKDB_EXCEL_URL_PREFIX)
    catalog_value, _separator, query = raw_target.partition("?")
    params = parse_qs(query, keep_blank_values=True)
    target_path = _normalize_url_path(catalog_value)
    workbook_path, catalog_path = _resolve_workbook_and_catalog(target_path, params)
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("DuckDB Excel workbook must be an .xlsx file.")
    if not workbook_path.is_file():
        raise ValueError(f"DuckDB Excel workbook does not exist: {workbook_path.name}.")

    mode = _single_param(params, "mode", DEFAULT_MODE)
    if mode not in {"table", "view"}:
        raise ValueError("DuckDB Excel mode must be either table or view.")

    sheets = _parse_sheet_mappings(params, workbook_path)
    validate_excel_config(workbook_path, sheets)
    catalog_path.parent.mkdir(parents=True, exist_ok=True)

    return DuckDBExcelConfig(
        catalog_path=catalog_path,
        workbook_path=workbook_path,
        sheets=sheets,
        always_read_current_file=mode == "view",
    )


def _resolve_workbook_and_catalog(
    target_path: str,
    params: dict[str, list[str]],
) -> tuple[Path, Path]:
    workbook_values = params.get("workbook") or []
    target_suffix = Path(target_path).suffix.lower()

    if target_suffix == ".xlsx":
        workbook_path = _resolve_path(target_path, "GAARD_DUCKDB_SOURCE_ROOT", "workbook")
        if workbook_values:
            raise ValueError(
                "DuckDB Excel URL should either point directly to an .xlsx file "
                "or use a .duckdb catalog with workbook=..., not both."
            )
        catalog_path = _default_catalog_path_for_workbook(workbook_path)
        return workbook_path, catalog_path

    if len(workbook_values) != 1:
        raise ValueError(
            "DuckDB Excel URL requires workbook=... when the URL points to a .duckdb catalog."
        )

    catalog_path = _resolve_path(target_path, "GAARD_DUCKDB_CATALOG_ROOT", "catalog")
    if catalog_path.name == ":memory:":
        raise ValueError("DuckDB Excel connector requires a persistent .duckdb catalog.")
    if catalog_path.suffix != ".duckdb":
        raise ValueError("DuckDB Excel catalog path must end with .duckdb or point to an .xlsx file.")

    workbook_path = _resolve_path(
        _normalize_url_path(workbook_values[0]),
        "GAARD_DUCKDB_SOURCE_ROOT",
        "workbook",
    )
    return workbook_path, catalog_path


def _default_catalog_path_for_workbook(workbook_path: Path) -> Path:
    catalog_root = os.getenv("GAARD_DUCKDB_CATALOG_ROOT")
    if catalog_root:
        catalog_path = Path(catalog_root).resolve() / f"{workbook_path.stem}.duckdb"
        _ensure_path_inside_root(catalog_path, "GAARD_DUCKDB_CATALOG_ROOT", "catalog")
        return catalog_path
    return workbook_path.with_suffix(".duckdb")


def _parse_sheet_mappings(
    params: dict[str, list[str]],
    workbook_path: Path,
) -> tuple[ExcelSheetMapping, ...]:
    explicit_sheets = params.get("sheet", [])
    if explicit_sheets:
        return tuple(_parse_sheet_mapping(value, params) for value in explicit_sheets)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    if not sheet_names:
        raise ValueError("DuckDB Excel workbook must contain at least one sheet.")

    used_relation_names: set[str] = set()
    mappings: list[ExcelSheetMapping] = []
    for sheet_name in sheet_names:
        relation_name = _unique_relation_name(_sanitize_relation_name(sheet_name), used_relation_names)
        mappings.append(
            ExcelSheetMapping(
                sheet_name=sheet_name,
                relation_name=relation_name,
                header=_bool_param(params, "header", True),
                range=_optional_param(params, "range"),
                all_varchar=_bool_param(params, "all_varchar", False),
                ignore_errors=_bool_param(params, "ignore_errors", False),
                type_sample_rows=_type_sample_rows_param(params),
            )
        )

    return tuple(mappings)


def _sanitize_relation_name(sheet_name: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9_]+", "_", sheet_name.strip())
    normalized = normalized.strip("_").lower()
    if not normalized or not re.match(r"^[a-zA-Z_]", normalized):
        normalized = f"sheet_{normalized}"
    normalized = normalized[:63]
    if normalized.lower() in RESERVED_RELATION_NAMES:
        normalized = f"{normalized}_sheet"
    return normalized


def _unique_relation_name(base_name: str, used_names: set[str]) -> str:
    relation_name = base_name
    suffix = 2
    while relation_name.lower() in used_names:
        suffix_text = f"_{suffix}"
        relation_name = f"{base_name[: 63 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(relation_name.lower())
    return relation_name


def validate_excel_config(workbook_path: Path, sheets: tuple[ExcelSheetMapping, ...]) -> None:
    relation_names = [sheet.relation_name for sheet in sheets]
    if len(relation_names) != len(set(relation_names)):
        raise ValueError("DuckDB Excel relation names must be unique.")

    for relation_name in relation_names:
        if not RELATION_NAME_PATTERN.match(relation_name):
            raise ValueError(f"Invalid DuckDB Excel relation name: {relation_name}.")
        if relation_name.lower() in RESERVED_RELATION_NAMES:
            raise ValueError(f"DuckDB Excel relation name is reserved: {relation_name}.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        existing_sheets = set(workbook.sheetnames)
    finally:
        workbook.close()

    missing = sorted({sheet.sheet_name for sheet in sheets} - existing_sheets)
    if missing:
        raise ValueError(f"DuckDB Excel workbook is missing sheet(s): {', '.join(missing)}.")


def provision_excel_relations(config: DuckDBExcelConfig) -> None:
    engine = create_engine(config.duckdb_url)
    try:
        try:
            with engine.begin() as connection:
                connection.execute(text("INSTALL excel"))
                connection.execute(text("LOAD excel"))
                _assert_read_xlsx_available(connection)
                _provision_relations(connection, config)
        except SQLAlchemyError as exc:
            raise ValueError(_friendly_duckdb_error(exc)) from exc
        except Exception as exc:
            raise ValueError(f"Excel File Loader provisioning failed: {exc}") from exc
    finally:
        engine.dispose()


def _assert_read_xlsx_available(connection: Connection) -> None:
    functions = connection.execute(
        text("SELECT function_name FROM duckdb_functions() WHERE function_name = 'read_xlsx'")
    ).fetchall()
    if not functions:
        raise ValueError(
            "DuckDB Excel extension is loaded, but read_xlsx is unavailable. "
            "Upgrade duckdb to 1.4.x in the GAARD environment and reinstall this connector."
        )


def _provision_relations(connection: Connection, config: DuckDBExcelConfig) -> None:
    temp_names = [f"__gaard_excel_{sheet.relation_name}" for sheet in config.sheets]
    for temp_name in temp_names:
        _drop_relation_if_exists(connection, temp_name)

    for sheet in config.sheets:
        temp_name = f"__gaard_excel_{sheet.relation_name}"
        read_sql = _read_xlsx_sql(config.workbook_path, sheet)
        if config.always_read_current_file:
            if sheet.type_sample_rows is not None:
                raw_name = f"__gaard_excel_raw_{sheet.relation_name}"
                _drop_relation_if_exists(connection, raw_name)
                connection.execute(text(f'CREATE TABLE "{raw_name}" AS {read_sql}'))
                typed_sql = _typed_select_sql(connection, config.workbook_path, sheet, raw_name, read_sql)
                connection.execute(text(f'CREATE OR REPLACE VIEW "{temp_name}" AS {typed_sql}'))
                _drop_relation_if_exists(connection, raw_name)
            else:
                connection.execute(text(f'CREATE OR REPLACE VIEW "{temp_name}" AS {read_sql}'))
        else:
            if sheet.type_sample_rows is not None:
                raw_name = f"__gaard_excel_raw_{sheet.relation_name}"
                _drop_relation_if_exists(connection, raw_name)
                connection.execute(text(f'CREATE TABLE "{raw_name}" AS {read_sql}'))
                typed_sql = _typed_select_sql(connection, config.workbook_path, sheet, raw_name)
                connection.execute(text(f'CREATE TABLE "{temp_name}" AS {typed_sql}'))
                _drop_relation_if_exists(connection, raw_name)
            else:
                connection.execute(text(f'CREATE TABLE "{temp_name}" AS {read_sql}'))
        connection.execute(text(f'DESCRIBE "{temp_name}"'))

    for sheet in config.sheets:
        relation_name = sheet.relation_name
        temp_name = f"__gaard_excel_{relation_name}"
        _drop_relation_if_exists(connection, relation_name)
        if config.always_read_current_file:
            final_sql = _read_xlsx_sql(config.workbook_path, sheet)
            if sheet.type_sample_rows is not None:
                final_sql = _typed_select_sql(connection, config.workbook_path, sheet, temp_name, final_sql)
            connection.execute(text(f'CREATE OR REPLACE VIEW "{relation_name}" AS {final_sql}'))
            _drop_relation_if_exists(connection, temp_name)
        else:
            connection.execute(text(f'ALTER TABLE "{temp_name}" RENAME TO "{relation_name}"'))


def _drop_relation_if_exists(connection: Connection, relation_name: str) -> None:
    row = connection.execute(
        text(
            """
            SELECT table_type
            FROM information_schema.tables
            WHERE table_schema = current_schema()
              AND table_name = :relation_name
            """
        ),
        {"relation_name": relation_name},
    ).fetchone()
    if row is None:
        return

    table_type = str(row[0]).upper()
    quoted_name = _quote_identifier(relation_name)
    if "VIEW" in table_type:
        connection.execute(text(f"DROP VIEW {quoted_name}"))
    else:
        connection.execute(text(f"DROP TABLE {quoted_name}"))


def _introspect_configured_relations(
    connection: Connection,
    config: DuckDBExcelConfig,
) -> DatabaseSchema:
    tables: list[TableInfo] = []
    object_type = "view" if config.always_read_current_file else "table"

    for sheet in config.sheets:
        rows = connection.execute(
            text(
                """
                SELECT column_name, data_type, is_nullable
                FROM information_schema.columns
                WHERE table_schema = current_schema()
                  AND table_name = :relation_name
                ORDER BY ordinal_position
                """
            ),
            {"relation_name": sheet.relation_name},
        ).fetchall()
        if not rows:
            raise ValueError(
                f"DuckDB Excel relation {sheet.relation_name!r} was not created successfully."
            )

        tables.append(
            TableInfo(
                name=sheet.relation_name,
                object_type=object_type,
                columns=[
                    ColumnInfo(
                        name=str(row[0]),
                        type=str(row[1]),
                        nullable=str(row[2]).upper() == "YES",
                        primary_key=False,
                    )
                    for row in rows
                ],
                foreign_keys=[],
            )
        )

    return DatabaseSchema(tables=tables)


def _read_xlsx_sql(workbook_path: Path, sheet: ExcelSheetMapping) -> str:
    effective_range = sheet.range or _detect_table_range(workbook_path, sheet)
    options: dict[str, Any] = {
        "sheet": sheet.sheet_name,
        "header": sheet.header,
        "all_varchar": sheet.all_varchar or sheet.type_sample_rows is not None,
        "ignore_errors": sheet.ignore_errors,
    }
    if effective_range:
        options["range"] = effective_range

    rendered_options = ", ".join(f"{key}={_duckdb_literal(value)}" for key, value in options.items())
    return f"SELECT * FROM read_xlsx({_duckdb_literal(str(workbook_path))}, {rendered_options})"


def _typed_select_sql(
    connection: Connection,
    workbook_path: Path,
    sheet: ExcelSheetMapping,
    raw_relation_name: str,
    live_read_sql: str | None = None,
) -> str:
    column_names = _relation_column_names(connection, raw_relation_name)
    inferred_types = _infer_duckdb_types_by_position(workbook_path, sheet, len(column_names))
    source_sql = f"SELECT * FROM {_quote_identifier(raw_relation_name)}"
    if live_read_sql is not None:
        source_sql = live_read_sql

    projection = ", ".join(
        (
            f"TRY_CAST({_quote_identifier(column_name)} AS {duckdb_type}) AS "
            f"{_quote_identifier(column_name)}"
        )
        if duckdb_type != "VARCHAR"
        else _quote_identifier(column_name)
        for column_name, duckdb_type in zip(column_names, inferred_types, strict=False)
    )
    return f"SELECT {projection} FROM ({source_sql})"


def _relation_column_names(connection: Connection, relation_name: str) -> list[str]:
    rows = connection.execute(
        text(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = current_schema()
              AND table_name = :relation_name
            ORDER BY ordinal_position
            """
        ),
        {"relation_name": relation_name},
    ).fetchall()
    return [str(row[0]) for row in rows]


def _infer_duckdb_types_by_position(
    workbook_path: Path,
    sheet: ExcelSheetMapping,
    column_count: int,
) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet.sheet_name]
        header_row = _detect_header_row(worksheet) if sheet.header else None
        min_column = 1
        max_column = column_count
        if header_row is not None:
            min_column, detected_max_column = _detect_header_columns(worksheet, header_row)
            max_column = min(detected_max_column, min_column + column_count - 1)
        min_row = (header_row + 1) if header_row is not None else None
        rows = worksheet.iter_rows(
            min_row=min_row,
            min_col=min_column,
            max_col=max_column,
            values_only=True,
        )
        if sheet.header:
            if header_row is None:
                next(rows, None)
        else:
            first_row = next(rows, None)
            if first_row is None:
                return ["VARCHAR"] * column_count
            rows = chain((first_row,), rows)

        samples = [[] for _column in range(column_count)]
        for row_index, row in enumerate(rows):
            if sheet.type_sample_rows is not None and row_index >= sheet.type_sample_rows:
                break
            for column_index, value in enumerate(row[:column_count]):
                samples[column_index].append(value)

        return [_infer_duckdb_type(values) for values in samples]
    finally:
        workbook.close()


def _detect_table_range(workbook_path: Path, sheet: ExcelSheetMapping) -> str | None:
    if not sheet.header:
        return None

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet.sheet_name]
        header_row = _detect_header_row(worksheet)
        if header_row is None or header_row == 1:
            return None
        min_column, max_column = _detect_header_columns(worksheet, header_row)
        max_row = max(header_row, worksheet.max_row or header_row)
        return f"{get_column_letter(min_column)}{header_row}:{get_column_letter(max_column)}{max_row}"
    finally:
        workbook.close()


def _detect_header_row(worksheet: Any) -> int | None:
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [value for value in row if value not in (None, "")]
        text_values = [value for value in values if isinstance(value, str) and value.strip()]
        if len(values) >= 2 and len(text_values) >= 2:
            return row_index
    return None


def _detect_header_columns(worksheet: Any, header_row: int) -> tuple[int, int]:
    row = next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
    non_empty_positions = [
        index
        for index, value in enumerate(row, start=1)
        if value not in (None, "") and str(value).strip()
    ]
    if not non_empty_positions:
        return 1, max(1, worksheet.max_column or 1)
    return min(non_empty_positions), max(non_empty_positions)


def _infer_duckdb_type(values: list[Any]) -> str:
    present_values = [value for value in values if value not in (None, "")]
    if not present_values:
        return "VARCHAR"
    if all(isinstance(value, bool) for value in present_values):
        return "BOOLEAN"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in present_values):
        return "BIGINT"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in present_values):
        return "DOUBLE"
    if all(isinstance(value, datetime) for value in present_values):
        return "TIMESTAMP"
    if all(isinstance(value, date) and not isinstance(value, datetime) for value in present_values):
        return "DATE"
    if all(isinstance(value, time) for value in present_values):
        return "TIME"
    return "VARCHAR"


def _quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _parse_sheet_mapping(value: str, params: dict[str, list[str]]) -> ExcelSheetMapping:
    if ":" not in value:
        raise ValueError("DuckDB Excel sheet mapping must use Sheet Name:relation_name.")

    sheet_name, relation_name = value.split(":", 1)
    if not sheet_name or not relation_name:
        raise ValueError("DuckDB Excel sheet mapping must include sheet and relation names.")

    return ExcelSheetMapping(
        sheet_name=sheet_name,
        relation_name=relation_name,
        header=_bool_param(params, "header", True),
        range=_optional_param(params, "range"),
        all_varchar=_bool_param(params, "all_varchar", False),
        ignore_errors=_bool_param(params, "ignore_errors", False),
        type_sample_rows=_type_sample_rows_param(params),
    )


def _resolve_path(path_value: str, root_env_name: str, label: str) -> Path:
    path = Path(path_value)
    if not path.is_absolute():
        raise ValueError(f"DuckDB Excel {label} path must be absolute.")

    resolved = path.resolve()
    _ensure_path_inside_root(resolved, root_env_name, label)
    return resolved


def _ensure_path_inside_root(path: Path, root_env_name: str, label: str) -> None:
    root_value = os.getenv(root_env_name)
    if root_value:
        root = Path(root_value).resolve()
        if path != root and root not in path.parents:
            raise ValueError(f"DuckDB Excel {label} path must be inside {root_env_name}.")


def _normalize_url_path(path_value: str) -> str:
    path_value = unquote(path_value)
    path_value = re.sub(r"\\+ ", " ", path_value)
    if sys.platform == "win32" and re.match(r"^/[a-zA-Z]:[\\/]", path_value):
        return path_value[1:]
    return path_value


def _single_param(params: dict[str, list[str]], key: str, default: str) -> str:
    values = params.get(key)
    if not values:
        return default
    if len(values) > 1:
        raise ValueError(f"DuckDB Excel URL accepts only one {key} parameter.")
    return values[0]


def _optional_param(params: dict[str, list[str]], key: str) -> str | None:
    value = _single_param(params, key, "")
    return value or None


def _bool_param(params: dict[str, list[str]], key: str, default: bool) -> bool:
    value = _single_param(params, key, str(default).lower()).lower()
    if value in {"1", "true", "yes"}:
        return True
    if value in {"0", "false", "no"}:
        return False
    raise ValueError(f"DuckDB Excel URL parameter {key} must be boolean.")


def _type_sample_rows_param(params: dict[str, list[str]]) -> int | None:
    value = _optional_param(params, "type_sample_rows") or _optional_param(params, "sample_size")
    if value is None:
        return None
    try:
        type_sample_rows = int(value)
    except ValueError as exc:
        raise ValueError("DuckDB Excel URL parameter type_sample_rows must be an integer.") from exc
    if type_sample_rows <= 0:
        raise ValueError("DuckDB Excel URL parameter type_sample_rows must be greater than 0.")
    return type_sample_rows


def _duckdb_literal(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def _friendly_duckdb_error(exc: SQLAlchemyError) -> str:
    detail = str(exc)
    if "read_xlsx" in detail and "does not exist" in detail:
        return (
            "Excel File Loader could not read XLSX files because DuckDB function read_xlsx "
            "is not available. "
            "The installed duckdb package is probably too old for the Excel adapter; "
            "install duckdb>=1.4,<1.5 and duckdb-engine>=0.17,<0.18 in the same venv, "
            "then restart GAARD."
        )
    if "INSTALL excel" in detail or "LOAD excel" in detail:
        return (
            "Excel File Loader could not install or load the Excel extension. "
            "Install the DuckDB Excel extension for this environment or allow DuckDB "
            "to download extensions during setup, then restart GAARD."
        )
    return f"Excel File Loader provisioning failed: {exc}"
